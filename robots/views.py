import json
import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from .models import Robot
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Count

@csrf_exempt
def create_robot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            serial = data.get('serial')
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')
            
            if not all([serial, model, version, created]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            if not Robot.objects.filter(model=model, version=version).exists():
                return JsonResponse({'error': f'Model {model} with version {version} does not exist'}, status=400)
            
            created_date = parse_datetime(created)
            if not created_date:
                return JsonResponse({'error': 'Invalid date format, expected YYYY-MM-DD HH:MM:SS'}, status=400)
            
            robot = Robot.objects.create(
                serial=serial,
                model=model,
                version=version,
                created=created_date
            )

            return JsonResponse({'message': 'Robot created successfully', 'robot_id': robot.id}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

    return JsonResponse({'error': 'Invalid method'}, status=405)


def generate_robot_report(request):
    last_week = datetime.datetime.now() - datetime.timedelta(days=7)

    robots_data = Robot.objects.filter(created__gte=last_week).values('model', 'version').annotate(count=Count('id')).order_by('model', 'version')

    wb = Workbook()

    models = set([robot['model'] for robot in robots_data])

    for model in models:

        sheet = wb.create_sheet(title=model)
        sheet.append(['Модель', 'Версия', 'Количество за неделю'])

        model_data = [robot for robot in robots_data if robot['model'] == model]
        
        for robot in model_data:
            sheet.append([robot['model'], robot['version'], robot['count']])
            
        for col in range(1, 4):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 20
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="robot_report.xlsx"'

    wb.save(response)

    return response




